"""
Twitter/X Bookmarks Extractor
Extrai tweets salvos (bookmarks) usando cookies do Chrome e exporta para Excel.
"""

import json
import time
import sys
import os
from datetime import datetime
from pathlib import Path

try:
    import browser_cookie3
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Dependências não encontradas. Execute: pip install -r requirements.txt")
    sys.exit(1)


TWITTER_API_BASE = "https://x.com/i/api/2"
GRAPHQL_ENDPOINT = "https://x.com/i/api/graphql"

# Query ID para bookmarks (pode mudar com atualizações do Twitter)
BOOKMARKS_QUERY_ID = "MhXAbUeE0RsRoCfv-bgscA"

BOOKMARKS_FEATURES = {
    "rweb_video_screen_enabled": False,
    "profile_label_improvements_pcf_label_in_post_enabled": True,
    "responsive_web_profile_redirect_enabled": False,
    "rweb_tipjar_consumption_enabled": False,
    "verified_phone_label_enabled": False,
    "creator_subscriptions_tweet_preview_api_enabled": True,
    "responsive_web_graphql_timeline_navigation_enabled": True,
    "responsive_web_graphql_skip_user_profile_image_extensions_enabled": False,
    "premium_content_api_read_enabled": False,
    "communities_web_enable_tweet_community_results_fetch": True,
    "c9s_tweet_anatomy_moderator_badge_enabled": True,
    "responsive_web_grok_analyze_button_fetch_trends_enabled": False,
    "responsive_web_grok_analyze_post_followups_enabled": True,
    "responsive_web_jetfuel_frame": True,
    "responsive_web_grok_share_attachment_enabled": True,
    "responsive_web_grok_annotations_enabled": True,
    "articles_preview_enabled": True,
    "responsive_web_edit_tweet_api_enabled": True,
    "graphql_is_translatable_rweb_tweet_is_translatable_enabled": True,
    "view_counts_everywhere_api_enabled": True,
    "longform_notetweets_consumption_enabled": True,
    "responsive_web_twitter_article_tweet_consumption_enabled": True,
    "tweet_awards_web_tipping_enabled": False,
    "responsive_web_grok_show_grok_translated_post": False,
    "responsive_web_grok_analysis_button_from_backend": True,
    "post_ctas_fetch_enabled": True,
    "freedom_of_speech_not_reach_fetch_enabled": True,
    "standardized_nudges_misinfo": True,
    "tweet_with_visibility_results_prefer_gql_limited_actions_policy_enabled": True,
    "longform_notetweets_rich_text_read_enabled": True,
    "longform_notetweets_inline_media_enabled": True,
    "responsive_web_grok_image_annotation_enabled": True,
    "responsive_web_grok_imagine_annotation_enabled": True,
    "responsive_web_grok_community_note_auto_translation_is_enabled": False,
    "responsive_web_enhance_cards_enabled": False,
}


def get_chrome_cookies():
    """Lê cookies exportados pelo Cookie-Editor (formato JSON)."""
    cookies_file = Path(__file__).parent / "cookies.json"
    print(f"Lendo cookies de: {cookies_file}")
    if not cookies_file.exists():
        print("ERRO: Arquivo cookies.json não encontrado.")
        print("Passos para exportar:")
        print("  1. Instale Cookie-Editor no Chrome: https://cookie-editor.com/")
        print("  2. Acesse x.com já logado")
        print("  3. Clique no ícone do Cookie-Editor")
        print("  4. Clique em Export > Export as JSON")
        print(f"  5. Salve como: {cookies_file}")
        sys.exit(1)
    try:
        with open(cookies_file, encoding="utf-8") as f:
            cookies_list = json.load(f)
        cookie_dict = {c["name"]: c["value"] for c in cookies_list}
        if not cookie_dict.get("auth_token"):
            print("ERRO: Cookie 'auth_token' não encontrado no arquivo.")
            print("Certifique-se de exportar os cookies estando logado no x.com.")
            sys.exit(1)
        print("Cookies lidos com sucesso. auth_token encontrado.")
        return cookie_dict
    except json.JSONDecodeError:
        print("ERRO: cookies.json não é um JSON válido.")
        sys.exit(1)
    except Exception as e:
        print(f"Erro ao ler cookies: {e}")
        sys.exit(1)


def get_guest_token(session):
    """Obtém o guest token necessário para algumas chamadas."""
    resp = session.post("https://api.twitter.com/1.1/guest/activate.json")
    if resp.status_code == 200:
        return resp.json().get("guest_token")
    return None


def build_session(cookies):
    """Cria uma sessão requests com os cookies e headers necessários."""
    session = requests.Session()

    # Definir cookies
    for name, value in cookies.items():
        session.cookies.set(name, value, domain=".x.com")

    ct0 = cookies.get("ct0", "")

    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Authorization": "Bearer AAAAAAAAAAAAAAAAAAAAANRILgAAAAAAnNwIzUejRCOuH5E6I8xnZz4puTs%3D1Zv7ttfk8LF81IUq16cHjhLTvJu4FA33AGWWjCpTnA",
        "x-csrf-token": ct0,
        "x-twitter-auth-type": "OAuth2Session",
        "x-twitter-client-language": "pt",
        "x-twitter-active-user": "yes",
        "content-type": "application/json",
        "Referer": "https://x.com/i/bookmarks",
    })

    return session


def parse_tweet(tweet_result):
    """Extrai dados relevantes de um resultado de tweet da API."""
    try:
        tweet = tweet_result.get("tweet", tweet_result)
        legacy = tweet.get("legacy", {})
        core = tweet.get("core", {})
        user_results = core.get("user_results", {}).get("result", {})
        user_core = user_results.get("core", {})
        user_legacy = user_results.get("legacy", {})

        # Texto completo (nota longa ou tweet normal)
        note_tweet = tweet.get("note_tweet", {})
        if note_tweet:
            full_text = note_tweet.get("note_tweet_results", {}).get("result", {}).get("text", legacy.get("full_text", ""))
        else:
            full_text = legacy.get("full_text", "")

        # Data
        created_at_str = legacy.get("created_at", "")
        created_at = ""
        if created_at_str:
            try:
                dt = datetime.strptime(created_at_str, "%a %b %d %H:%M:%S +0000 %Y")
                created_at = dt.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                created_at = created_at_str

        tweet_id = legacy.get("id_str", tweet.get("rest_id", ""))
        screen_name = user_core.get("screen_name", user_legacy.get("screen_name", ""))
        name = user_core.get("name", user_legacy.get("name", ""))
        url = f"https://x.com/{screen_name}/status/{tweet_id}" if screen_name and tweet_id else ""

        # Métricas
        retweets = legacy.get("retweet_count", 0)
        likes = legacy.get("favorite_count", 0)
        replies = legacy.get("reply_count", 0)
        quotes = legacy.get("quote_count", 0)
        views = tweet.get("views", {}).get("count", "")

        # Mídia
        media_list = legacy.get("extended_entities", {}).get("media", [])
        media_types = list(set(m.get("type", "") for m in media_list)) if media_list else []
        media_str = ", ".join(media_types) if media_types else ""

        # É retweet?
        is_retweet = "retweeted_status_result" in legacy
        is_quote = legacy.get("is_quote_status", False)

        return {
            "ID": tweet_id,
            "Data": created_at,
            "Autor": name,
            "@usuario": screen_name,
            "Texto": full_text,
            "Likes": likes,
            "Retweets": retweets,
            "Replies": replies,
            "Quotes": quotes,
            "Views": views,
            "Mídia": media_str,
            "É Retweet": "Sim" if is_retweet else "Não",
            "É Quote": "Sim" if is_quote else "Não",
            "URL": url,
        }
    except Exception as e:
        return None


def fetch_bookmarks(session, max_pages=50):
    """Busca todos os bookmarks via GraphQL API."""
    bookmarks = []
    cursor = None
    page = 0

    print("Iniciando extração de bookmarks...")

    while page < max_pages:
        variables = {
            "count": 100,
            "includePromotedContent": False,
        }
        if cursor:
            variables["cursor"] = cursor

        params = {
            "variables": json.dumps(variables),
            "features": json.dumps(BOOKMARKS_FEATURES),
        }

        url = f"{GRAPHQL_ENDPOINT}/{BOOKMARKS_QUERY_ID}/Bookmarks"

        try:
            resp = session.get(url, params=params, timeout=30)
        except Exception as e:
            print(f"Erro de conexão: {e}")
            break

        if resp.status_code == 401:
            print("ERRO 401: Não autorizado. Verifique se está logado no Chrome.")
            break
        elif resp.status_code == 429:
            print("Rate limit atingido. Aguardando 60 segundos...")
            time.sleep(60)
            continue
        elif resp.status_code != 200:
            print(f"ERRO HTTP {resp.status_code}: {resp.text[:200]}")
            break

        data = resp.json()

        # Navegar pela estrutura da resposta
        try:
            timeline = (
                data["data"]["bookmark_timeline_v2"]["timeline"]
            )
            instructions = timeline.get("instructions", [])
        except (KeyError, TypeError) as e:
            print(f"Estrutura de resposta inesperada: {e}")
            print(f"Resposta: {json.dumps(data)[:500]}")
            break

        new_cursor = None
        found_tweets = 0

        for instruction in instructions:
            if instruction.get("type") == "TimelineAddEntries":
                entries = instruction.get("entries", [])
                for entry in entries:
                    entry_id = entry.get("entryId", "")

                    # Cursor para próxima página
                    if "cursor-bottom" in entry_id or "cursor-top" in entry_id:
                        content = entry.get("content", {})
                        if content.get("entryType") == "TimelineTimelineCursor":
                            if "bottom" in entry_id:
                                new_cursor = content.get("value")
                        continue

                    # Tweet individual
                    content = entry.get("content", {})
                    item_content = content.get("itemContent", {})
                    tweet_result = item_content.get("tweet_results", {}).get("result", {})

                    if not tweet_result:
                        continue

                    # Debug: salva estrutura do primeiro tweet para diagnóstico
                    if not bookmarks and not found_tweets:
                        debug_path = Path(__file__).parent / "debug_tweet.json"
                        with open(debug_path, "w", encoding="utf-8") as dbf:
                            json.dump(tweet_result, dbf, indent=2, ensure_ascii=False)
                        print(f"[DEBUG] Estrutura do primeiro tweet salva em: {debug_path}")

                    parsed = parse_tweet(tweet_result)
                    if parsed:
                        bookmarks.append(parsed)
                        found_tweets += 1

        page += 1
        print(f"Página {page}: {found_tweets} tweets encontrados (total: {len(bookmarks)})")

        if not new_cursor or found_tweets == 0:
            print("Fim dos bookmarks.")
            break

        cursor = new_cursor
        time.sleep(1)  # Respeitar rate limits

    return bookmarks


def export_to_excel(bookmarks, output_path):
    """Exporta a lista de bookmarks para um arquivo Excel formatado."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookmarks"

    if not bookmarks:
        print("Nenhum bookmark para exportar.")
        return

    headers = list(bookmarks[0].keys())

    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="1DA1F2", end_color="1DA1F2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dados
    for row_idx, bookmark in enumerate(bookmarks, 2):
        for col_idx, header in enumerate(headers, 1):
            value = bookmark.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=(header == "Texto"), vertical="top")

        # Alternar cor de linha
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color="E8F5FE", end_color="E8F5FE", fill_type="solid"
                )

    # Ajustar largura das colunas
    col_widths = {
        "ID": 20, "Data": 20, "Autor": 20, "@usuario": 18,
        "Texto": 60, "Likes": 10, "Retweets": 12, "Replies": 10,
        "Quotes": 10, "Views": 10, "Mídia": 15,
        "É Retweet": 12, "É Quote": 10, "URL": 55,
    }
    for col_idx, header in enumerate(headers, 1):
        width = col_widths.get(header, 15)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Congelar linha de cabeçalho
    ws.freeze_panes = "A2"

    # Auto-filtro
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"\nArquivo salvo em: {output_path}")


def main():
    print("=" * 60)
    print("  Twitter/X Bookmarks Extractor")
    print("=" * 60)
    print()

    print("Certifique-se de ter o arquivo cookies.json na mesma pasta do script.")
    print("(Exportado pelo Cookie-Editor estando logado em x.com)")
    print()

    cookies = get_chrome_cookies()
    session = build_session(cookies)

    bookmarks = fetch_bookmarks(session)

    if not bookmarks:
        print("Nenhum bookmark encontrado.")
        return

    print(f"\nTotal de bookmarks extraídos: {len(bookmarks)}")

    # Nome do arquivo com timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = Path(__file__).parent
    output_path = output_dir / f"bookmarks_{timestamp}.xlsx"

    export_to_excel(bookmarks, str(output_path))
    print(f"Exportação concluída com sucesso!")
    print(f"Total: {len(bookmarks)} tweets salvos")


if __name__ == "__main__":
    main()
